import sys

VEHICLE_SPEED=1000
NUMBER_OF_TOTAL_CUSTOMERS_AND_DEPOT=276
NUMBER_OF_WORKING_DAYS_IN_A_WEEK=5
NUMBER_OF_WEEKS=2
NUMBER_OF_VEHICLES=1#per day
from openpyxl import Workbook
from openpyxl.utils import column_index_from_string, coordinate_from_string, get_column_letter
from ortools.constraint_solver import pywrapcp
import calendar
import openpyxl
import csv
import numpy
from geopy.distance import geodesic
import openpyxl
from sklearn.cluster import KMeans
import math
import datetime
import time

class XlsData:
    @classmethod
    def getXlsDataRange(cls,file, startCell, endCell, list_of_lists=False):
        wb = openpyxl.load_workbook(file)
        worksheet=wb.active
        rng = worksheet[startCell:endCell]#str:str
        data=[]
        if list_of_lists:
            for tuple in rng:
                row = ()
                for col in tuple:
                    row = row + (col.value,)
                data.append(list(row))
        else:
            for tuple in rng:
                row = ()
                for col in tuple:
                    row = row + (col.value,)
                data.append(row)
        return data

    @classmethod
    def getXlsDataStartOffset(cls,file, startCell, verticalOffset, horizontalOffset, list_of_lists=False):
        wb = openpyxl.load_workbook(file)
        worksheet=wb.active
        startCol = column_index_from_string((coordinate_from_string(startCell))[0])
        startRow = (coordinate_from_string(startCell))[1]
        endRow= startRow + verticalOffset
        endCol=startCol+horizontalOffset
        endCell=get_column_letter(endCol)+str(endRow)
        rng = worksheet[startCell:endCell]#str:str
        data=[]
        if list_of_lists:
            for tuple in rng:
                row = ()
                for col in tuple:
                    row = row + (col.value,)
                data.append(list(row))
        else:
            for tuple in rng:
                row = ()
                for col in tuple:
                    row = row + (col.value,)
                data.append(row)
        return data

    @classmethod
    def writeMatrixToXls(cls,filename,matrix):#might be useful to write matrices to file
        wb=Workbook()
        sheet = wb.active
        for i in range(len(matrix)):
            for x in range(len(matrix[i])):
                sheet.cell(row=i+1, column=x+1).value = matrix[i][x]
        wb.save(filename)
    @classmethod
    def iterateRowTillFirstValueCell(cls, startCell, file, targetValue=None):
        startCol=column_index_from_string((coordinate_from_string(startCell))[0])
        startRow = (coordinate_from_string(startCell))[1]
        wb = openpyxl.load_workbook(file)
        worksheet = wb.active
        row=[]
        #numcols=0
        for cell in worksheet.iter_cols(min_col=startCol, min_row=startRow, max_row=startRow):
            if cell[0].value == targetValue:
                break
            else:
                #numcols += 1
                row.append(tuple([float(i) for i in ((cell[0].value).split(')'))[1].split(';')]))
        '''lastcol=get_column_letter(startCol+numcols-1)
        lastcustomerdeparture=XlsData.getXlsDataRange(file, lastcol+str(startRow),lastcol+str(startRow), list_of_lists=False)
        lastcustomerdeparture = (lastcustomerdeparture[0])[0]
        lastcustomerdeparture =(lastcustomerdeparture.split(')'))[0]
        lastcustomerdeparture.replace("(", "")
        lastcustomerdeparture = (lastcustomerdeparture.split(';'))[1]
        row=(row,lastcustomerdeparture)'''
        return row
class DataInfo:
    @classmethod
    def computeRouteDistance(cls,listofcoords):
        totdist = 0
        for i in range(len(listofcoords)-1):
            totdist +=DataInfo.geodesic_distance_meters(listofcoords[i], listofcoords[i+1], as_integer=True)
        return totdist
    @classmethod
    def get_week_1_2_customers_indices(cls,all_data,num_weeks):
        one_week_frequence_customers_indices = [i for i in all_data.keys() if all_data[i]['frequence'] == 1]
        two_week_frequence_customers_indices = [i for i in all_data.keys() if
                                                all_data[i]['frequence'] == 2 and i != 0]  # excluding depot

        one_week_frequence_customers = [all_data[i]['gps_pos'] for i in all_data.keys() if
                                        all_data[i]['frequence'] == 1]
        week_clusterer = PositionClusterer(num_weeks, numpy.asarray([elem for elem in one_week_frequence_customers]),
                                           one_week_frequence_customers_indices)
        week_clusters = week_clusterer.compute_clustering()

        week1_customers_indices = two_week_frequence_customers_indices + week_clusters[0][1]
        week1_customers_indices.sort()
        week2_customers_indices = two_week_frequence_customers_indices + week_clusters[1][1]
        week2_customers_indices.sort()
        week_customer_indices={}
        week_customer_indices["week1"]=week1_customers_indices
        week_customer_indices["week2"] = week2_customers_indices
        return week_customer_indices

    @classmethod
    def get_day_clusters(cls, all_data, week1_customers_indices, week2_customers_indices):
        day_clusterer_week_1 = PositionClusterer(5, numpy.asarray(
            [all_data[i]['gps_pos'] for i in week1_customers_indices]),
                                                 week1_customers_indices)  # 5 is the number of days in a week
        day_clusterer_week_2 = PositionClusterer(5, numpy.asarray(
            [all_data[i]['gps_pos'] for i in week2_customers_indices]),
                                                 week2_customers_indices)  # 5 is the number of days in a week
        day_clusters_week_1 = day_clusterer_week_1.compute_clustering()
        day_clusters_week_2 = day_clusterer_week_2.compute_clustering()
        day_clusters = {}
        day_clusters[0] = day_clusters_week_1  # first week
        day_clusters[1] = day_clusters_week_2  # second week
        for i in day_clusters.keys():
            for j in day_clusters[i].keys():
                day_clusters[i][j][1].insert(0, 0)  # depot index
                day_clusters[i][j][1].sort()
        return day_clusters

    @classmethod
    def get_all_transportation_data(cls,pos_and_freq, tw_and_service_times, distances):
        all_data = {}
        for i in range(len(pos_and_freq)):
            posfreq = ((pos_and_freq[i][0], pos_and_freq[i][1]), pos_and_freq[i][2])
            twswk1 = (tw_and_service_times[i][0], tw_and_service_times[i][1], tw_and_service_times[i][2],
                      tw_and_service_times[i][3], tw_and_service_times[i][4])
            twewk1 = (tw_and_service_times[i][5], tw_and_service_times[i][6], tw_and_service_times[i][7],
                      tw_and_service_times[i][8], tw_and_service_times[i][9])
            twswk2 = (
                tw_and_service_times[i][10], tw_and_service_times[i][11], tw_and_service_times[i][12],
                tw_and_service_times[i][13],
                tw_and_service_times[i][14])
            twewk2 = (
                tw_and_service_times[i][15], tw_and_service_times[i][16], tw_and_service_times[i][17],
                tw_and_service_times[i][18],
                tw_and_service_times[i][19])
            sertime = tw_and_service_times[i][20]

            all_data[i] = {}
            all_data[i]['gps_pos'] = posfreq[0]
            all_data[i]['frequence'] = posfreq[1]
            all_data[i]['tw_wk1'] = [(twswk1[x], twewk1[x]) for x in range(len(twswk1))]
            all_data[i]['tw_wk2'] = [(twswk2[x], twewk2[x]) for x in range(len(twswk2))]
            all_data[i]['service_time'] = sertime
            all_data[i]['distances'] = distances[i]
        return all_data

    @classmethod
    def geodesic_distance_meters(cls,position_1, position_2, as_integer=False):#only works with gpsPositions
        """Computes the Manhattan distance between two points"""
        if as_integer:
            return int(geodesic(position_1,position_2).meters)
        else:
            return geodesic(position_1,position_2).meters

    @classmethod
    def geodesic_distance_matrix(cls,positions, as_integers=False):#may be useful to save the matrix then on file
        dist_mat=[]
        for  i in range(len(positions)):
            dist_mat.append([])
            for x in range(len(positions)):
                man = DataInfo.geodesic_distance_meters(positions[i], positions[x], as_integer=as_integers)
                dist_mat[i].insert(x,man)

        return dist_mat

    @classmethod
    def prepareRoutingData(cls, all_data, day_clusters, week, day):
        preproc={}
        preproc["coord"] = [all_data[z]['gps_pos'] for z in day_clusters[week][day][1]]
        tim = [None] * (len(day_clusters[week][day][1]))
        ser = [None] * (len(day_clusters[week][day][1]))
        if week == 0:
            for z in range(len(day_clusters[week][day][1])):
                tim[z] = (all_data[day_clusters[week][day][1][z]])['tw_wk1'][day]
        else:
            for z in range(len(day_clusters[week][day][1])):
                tim[z] = (all_data[day_clusters[week][day][1][z]])['tw_wk2'][day]
        preproc["time_windows"]=tim
        for z in range(len(day_clusters[week][day][1])):
            ser[z] = (all_data[day_clusters[week][day][1][z]])['service_time']
        preproc["service_times"]=ser
        dist = [None] * (len(day_clusters[week][day][1]))
        for z in range(len(day_clusters[week][day][1])):
            one_entry_dist = [None] * (len(day_clusters[week][day][1]))
            pos_ind = day_clusters[week][day][1][z]
            for x in range(len(day_clusters[week][day][1])):
                pos_ind2 = day_clusters[week][day][1][x]
                one_entry_dist[x] = all_data[pos_ind]['distances'][pos_ind2]
            dist[z] = one_entry_dist
        preproc["distances"]=dist
        preproc["global_indices"] = day_clusters[week][day][1]
        return preproc

    @classmethod
    def getPositionsFromIndices(cls, indices, all_data):
        return [all_data[i]['gps_pos'] for i in indices]

    @classmethod
    def getRoutesFromXls(cls, xlsfile,coords):
        depot=XlsData.getXlsDataRange(xlsfile, "A2", "A2", list_of_lists=False)
        depot=(depot[0])[0]
        depot.replace(" ","")
        depot = tuple([float(i) for i in depot.split(';')])
        route_coords_and_order = []
        for i in range(10):
            routeicustomerscoords = XlsData.iterateRowTillFirstValueCell(str("D"+str(i+3)), xlsfile, targetValue=None)
            routeicustomerscoords.insert(0, depot)
            routeicustomerscoords.append(depot)
            route_coords_and_order.append({})
            ord=[]
            dict={}
            for coord in routeicustomerscoords:
                ord.append(coords.index(coord))
            dict[0]=routeicustomerscoords
            dict[1]=ord
            route_coords_and_order[i]=dict
        return route_coords_and_order

    @classmethod
    def getRoutes(cls,xlsfile):
        depot = XlsData.getXlsDataRange(xlsfile, "A2", "A2", list_of_lists=False)
        depot = (depot[0])[0]
        depot.replace(" ", "")
        depot = tuple([float(i) for i in depot.split(';')])
        routeSummaries = []
        customercoords = []
        for i in range(10):
            departurei=XlsData.getXlsDataRange(xlsfile, str("B"+str(i+3)), str("B"+str(i+3)), list_of_lists=False)
            departurei = float((departurei[0])[0])
            arrivali = XlsData.getXlsDataRange(xlsfile, str("C" + str(i + 3)), str("C" + str(i + 3)),
                                                 list_of_lists=False)
            arrivali = float((arrivali[0])[0])
            routei = XlsData.iterateRowTillFirstValueCell(str("D"+str(i+3)), xlsfile, targetValue=None)

            customercoords = customercoords + [coord for coord in routei if coord not in customercoords]

            distanceRoute=DataInfo.computeRouteDistance([depot]+routei+[depot])
            routeDuration=arrivali-departurei
            routeSummary=[]
            routeSummary.append(distanceRoute)
            routeSummary.append(routeDuration)
            routeSummary.append(len(routei))
            routeSummaries.append(routeSummary)

        customercoords.insert(0, depot)
        return routeSummaries,customercoords

    @classmethod
    def getWeekDayNumber(cls, working_days_per_week,num_of_weeks, i):

        weeknumber = int(math.ceil((i + 1) / working_days_per_week))
        daynumber = int(i + 1 - (num_of_weeks * working_days_per_week / num_of_weeks) * (
            math.floor(i / working_days_per_week)))
        return weeknumber,daynumber
class PositionClusterer:
    def __init__(self, k,positions,position_indices):
        self.k=k#number of clusters
        self.positions=positions
        self.position_inidices=position_indices#len=len(positions)
    def compute_clustering(self):
        kmeans = KMeans(n_clusters=self.k).fit(self.positions)

        def getPositionsFromLabel(label,KMeansLabels,positions):
            return [(positions[i][0],positions[i][1]) for i in range(len(positions)) if KMeansLabels[i]==label]

        def getIndicesFromLabel(label,KMeansLabels,position_indices):
            return [position_indices[i] for i in range(len(KMeansLabels)) if KMeansLabels[i]==label]

        def aggregatePositionClusters(k):
            PositionClusters={}
            for i in range(k):
                PositionClusters[i]=(getPositionsFromLabel(i,kmeans.labels_,self.positions),getIndicesFromLabel(i,kmeans.labels_,self.position_inidices))
            return PositionClusters

        return aggregatePositionClusters(self.k)
class VrpSolver:
    def __init__(self, vehicle_speed,number_of_cust_and_depo,working_days_in_a_week,number_of_weeks,num_vehicles):
        self.vehicle_speed=vehicle_speed #meters per minute
        self.number_of_cust_and_depo=number_of_cust_and_depo
        self.working_days_in_a_week=working_days_in_a_week
        self.number_of_weeks=number_of_weeks
        self.num_vehicles=num_vehicles

    def computeRoutes(self,coordinates_file, distance_matrix_file):
        timestamp = datetime.datetime.fromtimestamp(time.time()).strftime('%Y-%m-%d_%H-%M-%S')
        filename = "route_vrp_" + timestamp
        pos_and_freq= XlsData.getXlsDataStartOffset(coordinates_file,"G20",self.number_of_cust_and_depo-1,2)#2=3-1, -1 beacuse G20 is already at the depot
                                                                                                        # gps pos and frequence including depot
        tw_and_service_times = XlsData.getXlsDataStartOffset(coordinates_file, "KW20",self.number_of_cust_and_depo-1,
                                                             self.working_days_in_a_week*self.number_of_weeks*2)  # time windows and service times including depot
        distances= XlsData.getXlsDataStartOffset(distance_matrix_file, "A1", self.number_of_cust_and_depo-1,self.number_of_cust_and_depo-1,
                                            list_of_lists=True)
        all_data = DataInfo.get_all_transportation_data(pos_and_freq, tw_and_service_times, distances)
        week_1_2_customer_indices = DataInfo.get_week_1_2_customers_indices(all_data,self.number_of_weeks)
        week1_customers_indices = week_1_2_customer_indices["week1"]  # excluding depot, that will be added later
        week2_customers_indices = week_1_2_customer_indices["week2"]

        day_clusters = DataInfo.get_day_clusters(all_data, week1_customers_indices, week2_customers_indices)
        routes = {}
        route_summaries = {}
        for i in day_clusters.keys():
            routes[i] = {}
            route_summaries[i] = {}
            for j in day_clusters[i].keys():
                preproc = DataInfo.prepareRoutingData(all_data, day_clusters, i, j)
                cus_cord_and_depot = preproc["coord"]
                time_windows = preproc["time_windows"]
                ser_times = preproc["service_times"]
                dist = preproc["distances"]
                global_indices = preproc["global_indices"]
                googlerouter = GoogleTWVRP(cus_cord_and_depot, self.num_vehicles, time_windows, ser_times, self.vehicle_speed, dist, global_indices,
                                           i,
                                           j, filename)
                result_routing_computation = googlerouter.compute_routing()
                route = result_routing_computation[0]
                if not route:
                    raise Exception('Error in Google route computer')
                route_summary = result_routing_computation[1]
                routes[i][j] = {}
                routes[i][j][1] = route
                route_summaries[i][j] = route_summary

        GoogleTWVRP.routingFromTxtToCsvXlsx(filename)
        for i in range(len(routes)):
            for j in range(len(routes[i])):
                positions = DataInfo.getPositionsFromIndices(routes[i][j][1], all_data)
                routes[i][j][0] = positions
        return routes, route_summaries, all_data,filename
class GoogleTWVRP:
    def __init__(self, customers_and_depot, num_vehicles, time_windows, service_times, speed, distances,global_indices,week,day,filename):
        self.locations = customers_and_depot
        self.num_vehicles= num_vehicles
        self.time_windows=time_windows
        self.service_times=service_times
        self.speed=speed
        self.distances=distances
        self.global_indices=global_indices
        self.week=week
        self.day=day
        self.filename=filename

    def getRoutingIndices(self,assignment, routing, global_indices):  # for one vehicle only
        index = routing.Start(0)
        routingIndices = []
        while not routing.IsEnd(index):
            node_index = routing.IndexToNode(index)
            routingIndices.append(global_indices[node_index])
            index = assignment.Value(routing.NextVar(index))

        node_index = routing.IndexToNode(index)
        routingIndices.append(global_indices[node_index])
        return routingIndices

    def create_data_model(self):
        """Stores the data for the problem"""
        data = {}

        data["locations"] = self.locations
        data["num_locations"] = len(data["locations"])
        data["num_vehicles"] = self.num_vehicles
        data["depot"] = 0
        data["time_windows"] = self.time_windows
        for i in range(0,len(data["time_windows"])):#including depot
            if data["time_windows"][i][1]==0:
                lst=list(data["time_windows"][i])
                lst[1]=3000
                data["time_windows"][i]=tuple(lst)
        data["service_times"]=self.service_times
        data["vehicle_speed"] = self.speed
        data["distances"]=self.distances
        return data

    #######################
    # Problem Constraints #
    #######################

    def create_distance_callback(self,data_distances):
        """Creates callback to return distance between points."""
        _distances = {}

        _distances=data_distances

        def distance_callback(from_node, to_node):
            return _distances[from_node][to_node]

        return distance_callback
    def create_time_callback(self,data):
        """Creates callback to get total times between locations."""

        def service_time(node):
            """Gets the service time for the specified location."""
            return data["service_times"][node]

        def travel_time(from_node, to_node):
            """Gets the travel times between two locations."""
            if from_node == to_node:
                travel_time = 0
            else:
                travel_time=(data["distances"][from_node][to_node])/data["vehicle_speed"]
            return travel_time

        def time_callback(from_node, to_node):
            """Returns the total time between the two nodes"""
            if from_node==to_node:
                return 0
            serv_time = service_time(from_node)
            trav_time = travel_time(from_node, to_node)
            return int(serv_time + trav_time)

        return time_callback

    def add_time_window_constraints(self,routing, data, time_callback):
        """Add Global Span constraint"""
        time = "Time"
        waiting_time_vehicle_horizon = 10

        max_min_tw=0
        try:
            for tw in data["time_windows"]:
                if tw[0] > max_min_tw:
                    max_min_tw = tw[0]
            max_time_vehicle_horizon = 540 + max_min_tw  # doesn't mean max time length to do all the route. It means the max time(minutes after midnight)
            # so in this case I say that it shouldn't take more than 9 hours after the time of the latest opening of the customers
        except Exception as ex:
            print(ex)
        routing.AddDimension(
            time_callback,
            waiting_time_vehicle_horizon,  # allow waiting time
            max_time_vehicle_horizon,  # maximum time per vehicle
            False,  # Don't force start cumul to zero. This doesn't have any effect here since the depot has a start window of (0, 0).
            time)
        time_dimension = routing.GetDimensionOrDie(time)
        for location_node, location_time_window in enumerate(data["time_windows"]):
            index = routing.NodeToIndex(location_node)
            try:
                time_dimension.CumulVar(index).SetRange(location_time_window[0], location_time_window[1])
            except Exception as ex:
                print(index<routing.nodes())
                print(time_dimension.CumulVar(index).Max()>=time_dimension.CumulVar(index).Min())
                print("time window error")
                print(ex)

    ###########
    # Printer #
    ###########
    def print_solution(self,data, routing, assignment,filename,local_positions,global_indices,week,day):
        """Prints assignment on console"""
        # Inspect solution.
        #capacity_dimension = routing.GetDimensionOrDie('Capacity')
        time_dimension = routing.GetDimensionOrDie('Time')
        total_dist = 0
        time_matrix = 0
        file_plan_output = 'Week {0}; day:{1}:\n'.format(week+1,calendar.day_name[day])
        for vehicle_id in range(data["num_vehicles"]):
            index = routing.Start(vehicle_id)
            plan_output = 'Route for vehicle {0}:\n'.format(vehicle_id)
            file_plan_output += 'Route for vehicle {0}:\n'.format(vehicle_id)
            route_dist = 0
            while not routing.IsEnd(index):
                node_index = routing.IndexToNode(index)
                next_node_index = routing.IndexToNode(
                    assignment.Value(routing.NextVar(index)))
                route_dist += data["distances"][node_index][next_node_index]
                #load_var = capacity_dimension.CumulVar(index)
                #route_load = assignment.Value(load_var)
                time_var = time_dimension.CumulVar(index)
                time_min = assignment.Min(time_var)
                time_max = assignment.Max(time_var)
                '''plan_output += ' {0} Load({1}) Time({2},{3}) ->'.format(
                    node_index,
                    route_load,
                    time_min, time_max)'''
                plan_output += ' {0} Time({1},{2}) ->'.format(
                    node_index,
                    time_min, time_max)
                all_data_index=global_indices[node_index]
                position=local_positions[node_index]
                file_plan_output+= ' {0} {1} Time({2},{3}) ->'.format(
                    all_data_index,position,
                    time_min, time_max)
                index = assignment.Value(routing.NextVar(index))

            node_index = routing.IndexToNode(index)
            #load_var = capacity_dimension.CumulVar(index)
            #route_load = assignment.Value(load_var)
            time_var = time_dimension.CumulVar(index)
            route_time = assignment.Value(time_var)
            time_min = assignment.Min(time_var)
            time_max = assignment.Max(time_var)
            total_dist += route_dist
            time_matrix += route_time
            '''plan_output += ' {0} Load({1}) Time({2},{3})\n'.format(node_index, route_load,
                                                                   time_min, time_max)'''
            plan_output += ' {0} Time({1},{2})\n'.format(node_index,
                                                                   time_min, time_max)
            plan_output += 'Distance of the route: {0} m\n'.format(route_dist)
            #plan_output += 'Load of the route: {0}\n'.format(route_load)
            plan_output += 'Time of the route: {0} min\n'.format(route_time)
            all_data_index = global_indices[node_index]
            position = local_positions[node_index]
            file_plan_output += ' {0} {1} Time({2},{3})\n'.format(all_data_index,position,
                                                         time_min, time_max)
            file_plan_output += 'Distance of the route: {0} m\n'.format(route_dist)
            # plan_output += 'Load of the route: {0}\n'.format(route_load)
            file_plan_output += 'Time of the route: {0} min\n'.format(route_time)
            file_plan_output += 'Number of customers: {0}\n'.format(data["num_locations"]-1)
            print(plan_output)
            with open(filename+".txt", "a") as text_file:
                print(file_plan_output, file=text_file)
            text_file.close()
        print('Total Distance of all routes: {0} m'.format(total_dist))
        print('Total Time of all routes: {0} min'.format(time_matrix))
        return [total_dist,time_matrix,data["num_locations"]-1]

    def compute_routing(self):
        """Entry point of the program"""
        # Instantiate the data problem.
        data = self.create_data_model()

        # Create Routing Model
        routing = pywrapcp.RoutingModel(data["num_locations"], data["num_vehicles"], data["depot"])
        # Define weight of each edge
        distance_callback = self.create_distance_callback(data["distances"])
        routing.SetArcCostEvaluatorOfAllVehicles(distance_callback)
        # Add Capacity constraint
        #demand_callback = self.create_demand_callback(data)
        #self.add_capacity_constraints(routing, data, demand_callback)
        # Add Time Window constraint
        time_callback = self.create_time_callback(data)
        self.add_time_window_constraints(routing, data, time_callback)

        # Setting first solution heuristic (cheapest addition).
        search_parameters = pywrapcp.RoutingModel.DefaultSearchParameters()
        #search_parameters.first_solution_strategy = (routing_enums_pb2.FirstSolutionStrategy.SAVINGS)
        #search_parameters.solution_limit = 1
        #search_parameters.optimization_step=15
        search_parameters.time_limit_ms = 20000
        '''search_parameters.local_search_metaheuristic = (
            routing_enums_pb2.LocalSearchMetaheuristic.SIMULATED_ANNEALING)'''
        # Solve the problem.
        try:
            assignment = routing.SolveWithParameters(search_parameters)
            if assignment:
                routesummary=self.print_solution(data, routing, assignment,self.filename,data["locations"],self.global_indices,self.week,self.day)
                routing_indices=self.getRoutingIndices(assignment,routing,self.global_indices)
                return routing_indices,routesummary
        except Exception as ex:
            print("solver fail")
            print(ex)
            return False

    @classmethod
    def routingFromTxtToCsvXlsx(cls,filename):
        with open(filename+'.txt', 'r') as in_file:
            stripped = (line.strip() for line in in_file)
            lines = (line.split("->") for line in stripped if line)
            with open(filename+'.csv', 'w') as out_file:
                writer = csv.writer(out_file)
                #writer.writerow(('title', 'intro'))
                writer.writerows(lines)
        wb = openpyxl.Workbook()
        ws = wb.active
        with open(filename + '.csv', 'r') as f:
            for row in csv.reader(f):
                ws.append(row)
        wb.save(filename + '.xlsx')
        in_file.close()
        out_file.close()
        f.close()
def main():
    coordinatesfile= sys.argv[1]
    distancesfile= sys.argv[2]
    vrpSolver = VrpSolver(VEHICLE_SPEED, NUMBER_OF_TOTAL_CUSTOMERS_AND_DEPOT, NUMBER_OF_WORKING_DAYS_IN_A_WEEK,
                          NUMBER_OF_WEEKS, NUMBER_OF_VEHICLES)
    routeComputation = vrpSolver.computeRoutes(coordinatesfile, distancesfile)


if __name__ == "__main__":
    main()
